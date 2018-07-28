"""
API for work with excel document.

Created on 13.06.2018

@author: Ruslan Dolovanyuk

"""

import logging
import os

import openpyxl

import xlrd
import xlwt
from xlutils.copy import copy as xlcopy

from configs import Config

from database import Database


class ApiXL:
    """Api class for work excel document."""

    def __init__(self):
        """Initialize ApiXL class."""
        self.log = logging.getLogger()
        self.xlsx = True
        self.rb = None
        self.wb = None
        self.rsheet = None
        self.sheet = None
        self.sheet_names = []
        self.xl_name = ''
        self.dir = ''

        self.db = Database()
        self.db.connect('settings.db')

        self.config = Config(self.db)

    def close(self):
        """Finish programm work."""
        self.db.disconnect()

    def open(self, file_name):
        """Open excel document."""
        self.xl_name = os.path.basename(file_name)
        self.xlsx = True if '.xlsx' == os.path.splitext(self.xl_name)[1] else False
        if self.xlsx:
            self.wb = openpyxl.load_workbook(file_name)
            self.sheet_names = self.wb.sheetnames
            self.sheet = self.wb[self.sheet_names[0]]
        else:
            try:
                self.rb = xlrd.open_workbook(file_name, on_demand=True, formatting_info=True)
                self.wb = xlcopy(self.rb)
            except:
                self.rb = xlrd.open_workbook(file_name, encoding_override='cp1251', on_demand=True, formatting_info=True)
                self.wb = xlcopy(self.rb)
            self.sheet_names = self.rb.sheet_names()
            self.rsheet = self.rb.get_sheet(0)
            self.sheet = self.wb.get_sheet(0)
        self.log.info('Файл %s открыт' % file_name)
        self.log.info('Редактирование %s' % self.sheet_names[0])

    def open_dir(self, dir_name):
        """Open dir with excel documents."""
        self.dir = dir_name
        self.log.info('Папка %s открыта для работы' % dir_name)

    def save(self):
        """Save excel document."""
        file_name = os.path.join(self.config.out_path, self.xl_name)
        self.wb.save(file_name)
        self.log.info('Файл %s сохранен' % file_name)

    def change_sheet(self, index):
        """Change sheet for editing."""
        if self.xlsx:
            self.sheet = self.wb[self.sheet_names[index]]
        else:
            self.rsheet = self.rb.get_sheet(index)
            self.sheet = self.wb.get_sheet(index)
        self.log.info('Редактирование %s' % self.sheet_names[index])

    def clear(self):
        """Close excel document."""
        self.rb = None
        self.wb = None
        self.rsheet = None
        self.sheet = None
        self.sheet_names.clear()
        self.xl_name = ''
        self.dir = ''

    def verify(self):
        """Clear some columns."""
        if self.wb is None:
            return
        for col in range(1, 3):
            self.clear_column(col)
        for col in range(8, 15):
            self.clear_column(col)

    def clear_column(self, col):
        """Clear one column."""
        col = col if self.xlsx else col-1
        row = 1 if self.xlsx else 0
        max_row = self.sheet.max_row if self.xlsx else self.rsheet.nrows
        while row < max_row:
            if self.xlsx:
                self.sheet.cell(row, col).value = ''
            else:
                self.sheet.write(row, col, xlrd.empty_cell.value)
            row += 1
        title = self.sheet.title if self.xlsx else self.rsheet.name
        self.log.info('Очищен %d столбец %s' % (col, title))

    def get_name_column(self, col):
        """Return name column."""
        name = openpyxl.get_column_letter(col) if self.xlsx else xlrd.colname(col-1)
        return name
